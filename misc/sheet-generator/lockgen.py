#!/usr/bin/python3
import openpyxl
import collections
import json
import fileDownloadfromGdrive
import boto3
from botocore.exceptions import NoCredentialsError

rootPath = ".."
localRootPath = "D:/Repositories/brutal"
#dev - uncomment the next line
#rootPath = localRootPath

#AWS S3 keys
with open(str(localRootPath) + "/misc/sheet-generator/s3-rootkey.json") as f:
	s3_dict=json.load(f)
ACCESS_KEY = s3_dict['AWSAccessKeyId']
SECRET_KEY = s3_dict['AWSSecretKey']

# Static Vars
totalNFLWeeks = 17


class Lock:
	def __init__(self, person, week, away, awayLine, home, homeLine, outcome):
		self.person = person
		self.week = (float(week[1:]) if str(week)[0] == '=' else week)
		self.away = away
		self.awayLine = (float(awayLine[1:]) if str(awayLine)[0] == '=' else awayLine)
		self.home = home
		self.homeLine = (float(homeLine[1:]) if str(homeLine)[0] == '=' else homeLine)
		self.outcome = outcome
	def __str__(self):
		return str(self.person) + " " + str(self.week) + " " + str(self.away) + " " + str(self.awayLine) + " " + str(self.home) + " " + str(self.homeLine) + " " + str(self.outcome)

def upload_to_aws(local_file, bucket, s3_file):
    s3 = boto3.client('s3', aws_access_key_id=ACCESS_KEY,
                      aws_secret_access_key=SECRET_KEY)

    try:
        s3.upload_file(local_file, bucket, s3_file, ExtraArgs={'ContentType': 'text/html'})
        print("Upload Successful")
        return True
    except FileNotFoundError:
        print("The file was not found")
        return False
    except NoCredentialsError:
        print("Credentials not available")
        return False

def write_top(file):
	file.write('<!DOCTYPE html>\n<html lang="en">\n\n<head>\n\t<!-- title -->\n\t<title>NFL Pickem | inov8.cc</title>\n\n\t<meta charset="utf-8">\n\t<meta http-equiv="X-UA-Compatible" content="IE=edge">\n\t<meta name="viewport" content="width=device-width, initial-scale=1">\n\n\t<!-- favicon -->\n\t<link href="' + str(rootPath) + '/assets/img/football.ico" rel="icon" type="image/x-icon" />\n\n\t<!-- jquery -->\n\t<script src="https://ajax.googleapis.com/ajax/libs/jquery/1.12.0/jquery.min.js"></script>\n\n\t<!-- BootStrap -->\n\t<link rel="stylesheet" type="text/css" href="http://netdna.bootstrapcdn.com/bootstrap/3.0.2/css/bootstrap.min.css">\n\t<script src="http://netdna.bootstrapcdn.com/bootstrap/3.0.2/js/bootstrap.min.js"></script>\n\n\t<!--CSS-->\n\t<link rel="stylesheet" type="text/css" href="' + str(rootPath) + '/assets/css/pickem.css">\n</head>\n\n<body>\n\t<nav class="menu site-nav"><a class="site-title" href="../index.html">inov8.cc</a>\n\t\t<div class="main-nav">\n\t\t\t<ul>\n\t\t\t\t<li>Pages</li>|\n\t\t\t\t<li><a href="../index.html">Home</a></li>\n\t\t\t\t<li><a href="../documents.html">Documents</a></li>\n\t\t\t</ul>\n\t\t</div>\n\t</nav>\n\n')
	file.write('\t<div class="container">\n')
	file.write('\t\t<div class="row">\n\n')

def write_comment(file, comment_text):
		file.write('\t\t\t<!-- ' + comment_text + '-->\n')

def write_person_opener(file, name, wins, losses, ties):
		file.write('\t\t\t<div class="col-xs-4">\n\t\t\t\t<div class="row">\n\t\t\t\t\t<div class="col-xs-10 col-xs-offset-1 col-sm-8 col-sm-offset-2 col-md-6 col-md-offset-3">\n\t\t\t\t\t\t<div class="relative-container">\n\t\t\t\t\t\t\t<img src="' + str(rootPath) + '/assets/img/' + str(name) + '.jpg" alt="' + str(name) + '" class="img-responsive img-circle">\n\t\t\t\t\t\t\t<div class="top-left">')
		if name == 'jack':
			file.write('\n\t\t\t\t\t\t\t\t<p><strong>Champion</strong></p>\n\t\t\t\t\t\t\t\t<p>2017</p>\n\t\t\t\t\t\t\t\t<p>2018</p>\n\t\t\t\t\t\t\t\t<p>2019</p>\n\t\t\t\t\t\t\t\t<p>2021</p>\n\t\t\t\t\t\t\t')
		elif name == 'kenny':
			file.write('\n\t\t\t\t\t\t\t\t<p><strong>Champion</strong></p>\n\t\t\t\t\t\t\t\t<p>2020</p>\n\t\t\t\t\t\t\t')
		elif name == 'eric':
			file.write('\n\t\t\t\t\t\t\t\t<p><strong>Champion</strong></p>\n\t\t\t\t\t\t\t\t<p>2022</p>\n\t\t\t\t\t\t\t')			
		file.write('</div>\n\t\t\t\t\t\t</div>\n\t\t\t\t\t</div>\n\t\t\t\t</div>\n\t\t\t\t<h4 class="text-center">' + str(wins) + '-' + str(losses) + '-' + str(ties) + '</h4>\n\t\t\t\t<table class="table">\n\t\t\t\t\t<thead>\n\t\t\t\t\t\t<tr class="hidden-xs">\n\t\t\t\t\t\t\t<th>Away</th>\n\t\t\t\t\t\t\t<th></th>\n\t\t\t\t\t\t\t<th>Home</th>\n\t\t\t\t\t\t\t<th></th>\n\t\t\t\t\t\t</tr>\n\t\t\t\t\t</thead>\n\t\t\t\t\t<tbody>\n')

def open_week(file, person, week, wins, losses, ties, kings, dks, isCurrentWeek):
	file.write('\t\t\t\t\t\t<tr><td colspan="4"><h4 class="text-center"><a data-toggle="collapse" href=".week' + str((week+1)) + str(person) +  '">' + ('<img src="' + str(rootPath) if kings > 0 else '') + ('/assets/img/crown.png" class="dk-icon" />' if kings > 0 else '') +'Wk ' +  str(week+1) + ('' if isCurrentWeek else '<span class="hidden-xs">: ' + str(wins) + '-' + str(losses) + '-' + str(ties) + '</span>') + ('<img src="' + str(rootPath) if dks > 0 else '') + ('/assets/img/dk.png" class="dk-icon" />' if dks > 0 else '') + '</a></h4></td></tr>\n\n')

def write_lock(file, lock, isCurrentWeek):
	print(lock)
	file.write('\t\t\t\t\t\t<tr class="' + ('' if isCurrentWeek else 'collapse') + ' week' + str(int(lock.week)) + str(lock.person)) 
	if lock.outcome != None:
		if 'C' in lock.outcome.upper():
			file.write(' success')
		elif 'P' in lock.outcome.upper():
			file.write(' warning')
		# elif 'S' in lock.outcome.upper() and isCurrentWeek:
			# file.write(' danger')
	file.write('">\n')
	if str(lock.awayLine) != "None":
		file.write('\t\t\t\t\t\t\t<td class="pick">' + str(lock.away) + '</td>\n')
		file.write('\t\t\t\t\t\t\t<td>' + (("+"+str(lock.awayLine)) if  (lock.awayLine >= 0) else str(lock.awayLine)) + '</td>\n')
	else:
		file.write('\t\t\t\t\t\t\t<td class="hidden-xs">' + str(lock.away) + '</td>\n')
		file.write('\t\t\t\t\t\t\t<td class="hidden-xs"></td>\n')
	if str(lock.homeLine) != "None":
		file.write('\t\t\t\t\t\t\t<td class="pick">' + str(lock.home) + '</td>\n')
		file.write('\t\t\t\t\t\t\t<td>' + (("+"+str(lock.homeLine)) if  (lock.homeLine >= 0) else str(lock.homeLine)) + '</td>\n')
	else:
		file.write('\t\t\t\t\t\t\t<td class="hidden-xs">' + str(lock.home) + '</td>\n')
		file.write('\t\t\t\t\t\t\t<td class="hidden-xs"></td>\n')
	file.write('\t\t\t\t\t\t</tr>\n')

def write_person_bottom(file):
	file.write('\t\t\t\t\t</tbody>\n\t\t\t\t</table>\n\t\t\t</div>\n\n')

def close_body(file):
	file.write('\t\t</div><!-- end row -->\n\n')
	file.write('\t\t<br />\n')
	file.write('\t\t<br />\n')
	file.write('\t\t<br />\n')
	file.write('\t\t<br />\n')
	file.write('\t\t<br />\n')
	file.write('\t\t<div class="row">\n')
	file.write('\t\t\t<div class="col-xs-12">\n')

	file.write('\t\t\t\t<p><strong>Locks: </strong> $250</p>\n')
	file.write('\t\t\t\t<p><strong>Survivor: </strong> $75</p>\n')
	file.write('\t\t\t</div>\n')
	file.write('\t\t</div><!-- end row -->\n')
	file.write('\t</div><!-- end container -->\n')
	file.write('</body>\n\n</html>')

#----------------------------------------------------------------------
def parse_file(path):
	book = openpyxl.load_workbook(path)
	outfile = open(str(localRootPath) + "/documents/pickem-2023.html", 'w')

	#write head, nav and body opening tag
	write_top(outfile)

	highestWeekGlobal = 0

	#iterate through sheets
	for s in range(len(book.sheetnames)):
		book.active = s
		sheet = book.active
		weeks = {x: [] for x in range(18)}

		#Iterate through rows and read data
		for i, row in enumerate(sheet.iter_rows()):                              
			if i == 0:                                                                              
				continue
			elif row[0].value is None:
				break                                                             
			lock = Lock(book.sheetnames[s], row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)
			weeks[(row[0].value - 1)].append(lock)

		#write name comment
		write_comment(outfile, book.sheetnames[s])

		#write div opener for each person
		#calculate total record
		totalWins = 0
		totalLosses = 0
		totalTies = 0
		highestWeekLocal = 0

		#determine current week
		for j in sorted (weeks.keys()):
			if len(weeks[j]) > 0:
				if j > highestWeekGlobal:
					highestWeekGlobal = j
				if j > highestWeekLocal:
					highestWeekLocal = j

		for j in sorted (weeks.keys()):
			if len(weeks[j]) > 0 and (j < highestWeekGlobal or j == totalNFLWeeks):
				
				#calculate record
				for k in weeks[j]:
					if k.outcome == None:
						totalLosses = totalLosses + 1
					elif 'C' in k.outcome.upper():
						totalWins = totalWins + 1
					elif 'P' in k.outcome.upper():
						totalTies = totalTies + 1
					else:
						totalLosses = totalLosses + 1


		write_person_opener(outfile, book.sheetnames[s], totalWins, totalLosses, totalTies)

		# calculate week record
		for j in sorted (weeks.keys()):
			if len(weeks[j]) > 0:

				#calculate record
				wins = 0
				losses = 0
				ties = 0
				kings = 0
				dks = 0
				for k in weeks[j]:  				
					if k.outcome == None:
						losses = losses + 1
					else:						
						if 'C' in k.outcome.upper():
							wins = wins + 1
						elif 'P' in k.outcome.upper():
							ties = ties + 1
						else:
							losses = losses + 1

						# look for king wins
						if 'K' in k.outcome.upper():
							kings = kings + 1
						
						# look for DK wins
						if 'D' in k.outcome.upper():
							dks = dks + 1
				
				#determine if week is current week
				isCurrentWeek = (j == highestWeekGlobal and j < totalNFLWeeks)
				
				#write week header
				open_week(outfile, book.sheetnames[s], j, wins, losses, ties, kings, dks, isCurrentWeek)

				for k in weeks[j]:  
					write_lock(outfile, k, isCurrentWeek)
		
		#close out each persons tbody, table, and div
		write_person_bottom(outfile)
	
	#write closing html tag
	close_body(outfile)
 
#----------------------------------------------------------------------
if __name__ == "__main__":
	print('starting...')
	fileDownloadfromGdrive.execute()
	print('successfully downloaded from g drive...')
	path = str(localRootPath) + "/misc/sheet-generator/pickem-from-gsheet.xlsx"
	parse_file(path)

	print('attempting to upload to aws...')
	fileToUpload = str(localRootPath) + "/documents/pickem-2023.html"
	uploaded = upload_to_aws(fileToUpload, 'inov8.cc', 'documents/pickem-2023.html')
	print('upload status: ' + str(uploaded))